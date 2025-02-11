#Task 1
import pandas as pd
import numpy as np
import csv as cs

data = {
    'student_id': [101, 102, 103, 104,105],
    'Name': ['James','Francise','Albino','Dante','Zade'],
    'Age':[19,20,19,18,21]
}

df = pd.DataFrame(data)
df.to_csv('C:\\Users\\Dell\\Documents\\AI.csv', index=False)

df1=pd.read_csv('C:\\Users\\Dell\\Documents\\AI.csv')

print(df1)



#Task 2

sheets = pd.read_excel("F:\\New folder\\Second.xlsx", sheet_name=None)
df3 = sheets['Sheet1']
df4 = sheets['Sheet2']

print(df3.head())
print(df4.head())

# Task 3

data = {
    'Day': ['Monday', 'Monday', 'Tuesday', 'Tuesday', 'Wednesday', 'Wednesday', 'Thursday', 'Thursday', 'Friday', 'Friday'],
    'Course': ['Mathematics', 'Physics', 'Chemistry', 'Biology', 'Computer Science', 'History', 'Economics', 'Statistics', 'English', 'Philosophy'],
    'Time': ['9:00 AM - 10:00 AM', '10:15 AM - 11:15 AM', '9:00 AM - 10:00 AM', '10:15 AM - 11:15 AM', 
             '9:00 AM - 10:00 AM', '10:15 AM - 11:15 AM', '9:00 AM - 10:00 AM', '10:15 AM - 11:15 AM', 
             '9:00 AM - 10:00 AM', '10:15 AM - 11:15 AM']
}

timetable=pd.DataFrame(data)
print("Time Table")
print(timetable)

#Task 4
import pandas as pd

# Task 4
df5 = pd.read_excel(r"C:\Users\Dell\Documents\AI\2023-CS-649.xlsx", usecols=['Roll', 'SubjectA','SubjectB','SubjectC','SubjectD'])
marks = df5[(df5['SubjectA'] > 80) | (df5['SubjectB'] > 80) | 
            (df5['SubjectC'] > 80) | (df5['SubjectD'] > 80)]
if not marks.empty:
    print(marks)
else:
    print("No marks greater than 80.")

# Task 5
df7=pd.read_excel(r"C:\Users\Dell\Documents\AI\2023-CS-649.xlsx", usecols=['Roll', 'SubjectA','SubjectB','SubjectC','SubjectD'])
grading_scale = {
    range(90, 101): 10,
    range(80, 90): 9,
    range(70, 80): 8,
    range(60, 70): 7,
    range(50, 60): 6,
    range(40, 50): 5,
    range(33, 40): 4,
    range(20, 33): 3,
    range(0, 20): 0
}

def get_grade_point(percentage):
    for score_range, grade_point in grading_scale.items():
        if percentage in score_range:
            return grade_point

df7['SubjectA_GP'] = df7['SubjectA'].apply(get_grade_point)
df7['SubjectB_GP'] = df7['SubjectB'].apply(get_grade_point)
df7['SubjectC_GP'] = df7['SubjectC'].apply(get_grade_point)
df7['SubjectD_GP'] = df7['SubjectD'].apply(get_grade_point)

df7['GPA'] = (df7['SubjectA_GP'] + df7['SubjectB_GP'] + df7['SubjectC_GP'] + df7['SubjectD_GP']) / 4

print(df7)



#Pandas Code

import pandas as pd
# Create a DataFrame
data = {
'Column1': [1, 2, 3],
'Column2': ['A', 'B', 'C']
}
df = pd.DataFrame(data)
# Write the DataFrame to an Excel file
df.to_excel("F:\\New folder\\Second.xlsx", index=False)

df=pd.read_excel("F:\\New folder\\Second.xlsx")
print(df.head())

sheets=pd.read_excel("F:\\New folder\\Second.xlsx",sheet_name=None)
df1=sheets['Sheet1']
df2=sheets['Sheet2']

print(df1.head())
print(df2.head())

with pd.ExcelWriter("F:\\New folder\\Second.xlsx") as writer:
    df1.to_excel(writer,sheet_name='Sheet1',index=False)
    df2.to_excel(writer,sheet_name='Sheet2',index=False)

df=pd.read_excel("F:\\New folder\\Second.xlsx",usecols=['Column1','Column2'])

df=pd.read_excel("F:\\New folder\\Second.xlsx",skiprows=2)
print(df.head())

from openpyxl import Workbook

wb=Workbook()
ws=wb.active
ws.title='Sheet1'

ws['A1']='Hello'
ws['B1']='World'

wb.save("F:\\New folder\\Second.xlsx")


# numpy codes

import pandas as pd
data_array = df.to_numpy()
import numpy as np
mean_values = np.mean(data_array, axis=0) 
print(mean_values)


data_array = np.array([[1, 2, 3], [4, 5, 6], [7, 8, 9]])
# Convert the NumPy array to a DataFrame
df = pd.DataFrame(data_array, columns=['Column1', 'Column2', 'Column3'])
# Write the DataFrame to an Excel file
df.to_excel('path_to_your_file.xlsx', sheet_name='Sheet1', index=False)

df = pd.read_excel('path_to_your_file.xlsx', sheet_name='Sheet1')
# Convert to a NumPy array
data_array = df.to_numpy()
std_devs = np.std(data_array, axis=0)
print(std_devs)


